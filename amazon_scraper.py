import requests
import pandas as pd
import re
import json
import time
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import logging
from config import AUTH_TOKEN

class AmazonScraper:
    def __init__(self, input_file='input.xlsx', output_file='output.xlsx'):
        self.input_file = input_file
        self.output_file = output_file
        self.max_retries = 3
        self.setup_logging()
        
        # Headers for the API (using the same from Google Sheets script)
        self.headers = {
            "accept": "application/json",
            "content-type": "application/json",
            "authorization": "Basic {AUTH_TOKEN}"
        }
        
        # API endpoint from Google Sheets script
        self.api_url = "https://scraper-api.decodo.com/v2/scrape"
        
        print("🚀 Amazon Scraper initialized successfully!")
        print(f"📁 Input file: {self.input_file}")
        print(f"📁 Output file: {self.output_file}")
    
    def setup_logging(self):
        """Setup logging configuration"""
        # Create a custom formatter that removes emojis for file logging
        class NoEmojiFormatter(logging.Formatter):
            def format(self, record):
                # Remove common emojis from log messages for file compatibility
                message = super().format(record)
                emoji_map = {
                    '🔄': '[RETRY]',
                    '✅': '[SUCCESS]',
                    '❌': '[ERROR]',
                    '⚠️': '[WARNING]',
                    '🔍': '[PROCESSING]',
                    '🎉': '[COMPLETED]',
                    '📊': '[INFO]',
                    '🚀': '[START]',
                    '📁': '[FILE]',
                    '📄': '[LOG]'
                }
                for emoji, replacement in emoji_map.items():
                    message = message.replace(emoji, replacement)
                return message
        
        # Setup logging with UTF-8 encoding and custom formatter
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('scraper.log', encoding='utf-8'),
                logging.StreamHandler()
            ]
        )
        
        # Apply no-emoji formatter to file handler only
        logger = logging.getLogger(__name__)
        for handler in logger.handlers:
            if isinstance(handler, logging.FileHandler):
                handler.setFormatter(NoEmojiFormatter('%(asctime)s - %(levelname)s - %(message)s'))
        
        self.logger = logger
    
    def validate_input_file(self):
        """Validate if input file exists and has correct structure"""
        if not os.path.exists(self.input_file):
            raise FileNotFoundError(f"❌ Input file '{self.input_file}' not found!")
        
        try:
            df = pd.read_excel(self.input_file)
            required_columns = ['PRODUCT_URL']
            
            if df.empty:
                raise ValueError("❌ Input file is empty!")
            
            # Check if PRODUCT_URL column exists (case insensitive)
            url_column = None
            for col in df.columns:
                if 'url' in str(col).lower():
                    url_column = col
                    break
            
            if url_column is None:
                raise ValueError("❌ No URL column found! Please ensure your Excel file has a column containing 'URL' in its name.")
            
            print(f"✅ Input file validated successfully!")
            print(f"📊 Found {len(df)} rows to process")
            return df, url_column
            
        except Exception as e:
            raise ValueError(f"❌ Error reading input file: {str(e)}")
    
    def create_output_file(self, df, url_column):
        """Create output file with proper headers"""
        # Create output dataframe with all necessary columns
        output_df = pd.DataFrame()
        output_df['PRODUCT_URL'] = df[url_column]
        output_df['Availability'] = ''
        output_df['Price'] = ''
        output_df['No_of_Reviews'] = ''
        output_df['Rating'] = ''
        output_df['Electronics_Rank'] = ''
        output_df['Last_Updated'] = ''
        output_df['Status'] = 'Pending'
        
        # Save to Excel
        output_df.to_excel(self.output_file, index=False)
        print(f"✅ Output file '{self.output_file}' created successfully!")
        return output_df
    
    def scrape_single_product(self, url):
        """Scrape a single product with retry logic"""
        for attempt in range(1, self.max_retries + 1):
            try:
                # Clean and validate URL
                url = str(url).strip()
                if not url.startswith("http"):
                    url = "https://" + url
                
                self.logger.info(f"🔄 Attempting to scrape (Attempt {attempt}/{self.max_retries}): {url}")
                
                # Prepare API request
                payload = {"url": url}
                options = {
                    'headers': self.headers,
                    'json': payload,
                    'timeout': 30
                }
                
                # Make API request
                response = requests.post(self.api_url, **options)
                
                if response.status_code == 200:
                    json_response = response.json()
                    html = json_response.get('results', [{}])[0].get('content', '')
                    
                    if not html:
                        if attempt < self.max_retries:
                            self.logger.warning(f"⚠️ Empty HTML content, retrying in 2 seconds...")
                            #time.sleep(2)
                            continue
                        else:
                            return self.create_error_result("Empty HTML content")
                    
                    # Extract data using the same logic as Google Sheets script
                    data = self.extract_data_from_html(html, url)
                    
                    # Check if we got valid data (price extraction check like in Google Sheets)
                    if data['availability'] == "In Stock" and data['price'] == "N/A" and attempt < self.max_retries:
                        self.logger.warning(f"⚠️ In stock but no price found, retrying in 2 seconds...")
                        #time.sleep(2)
                        continue
                    
                    self.logger.info(f"✅ Successfully scraped: {url}")
                    return data
                
                else:
                    self.logger.error(f"❌ API request failed with status {response.status_code}")
                    if attempt < self.max_retries:
                        #time.sleep(2)
                        continue
                        
            except Exception as e:
                self.logger.error(f"❌ Attempt {attempt} failed: {str(e)}")
                if attempt < self.max_retries:
                    #time.sleep(2)
                    continue
        
        return self.create_error_result("All retry attempts failed")
    
    def extract_data_from_html(self, html, url):
        """Extract product data from HTML using the same regex patterns as Google Sheets script"""
        data = {
            'url': url,
            'availability': 'N/A',
            'price': 'N/A',
            'reviews': '0',
            'rating': '0',
            'electronics_rank': 'N/A',
            'last_updated': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'status': 'Success'
        }
        
        try:
            # Extract availability (same logic as Google Sheets)
            if 'id="outOfStock"' in html:
                data['availability'] = "Out of Stock"
            else:
                data['availability'] = "In Stock"
            
            # Extract price from PPD section
            ppd_match = re.search(r'<div[^>]*id="ppd"[^>]*>([\s\S]*?)<\/div>\s*<\/div>\s*<\/div>', html)
            if ppd_match:
                ppd_html = ppd_match.group(1)
                price_match = re.search(r'class="a-price-whole">(.*?)<', ppd_html)
                if price_match:
                    data['price'] = price_match.group(1).strip().replace(',', '')
            
            # Extract reviews and rating from ACR section
            acr_match = re.search(r'<div[^>]*id="averageCustomerReviews"[^>]*>([\s\S]*?)<\/div>\s*<\/div>', html)
            if acr_match:
                acr_html = acr_match.group(1)
                
                # Extract review count
                reviews_match = re.search(r'id="acrCustomerReviewText"[^>]*>([\d,]+)\s+ratings?<', acr_html)
                if reviews_match:
                    data['reviews'] = reviews_match.group(1).replace(',', '')
                
                # Extract rating
                rating_match = re.search(r'class="a-icon-alt">([\d.]+) out of 5', acr_html)
                if rating_match:
                    data['rating'] = rating_match.group(1)
            
            # Extract electronics rank
            rank_match = re.search(r'#([\d,]+)\s+in\s+Electronics', html, re.IGNORECASE)
            if rank_match:
                data['electronics_rank'] = rank_match.group(1).replace(',', '')
                
        except Exception as e:
            self.logger.error(f"❌ Error extracting data: {str(e)}")
            data['status'] = f'Extraction Error: {str(e)}'
        
        return data
    
    def create_error_result(self, error_message):
        """Create error result structure"""
        return {
            'url': '',
            'availability': 'Error',
            'price': 'N/A',
            'reviews': '0',
            'rating': '0',
            'electronics_rank': 'N/A',
            'last_updated': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'status': error_message
        }
    
    def update_output_file(self, row_index, data):
        """Update output file with scraped data"""
        try:
            # Load the workbook
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            # Update the row (row_index + 2 because of header and 0-based index)
            row = row_index + 2
            
            ws.cell(row=row, column=2, value=data['availability'])  # Availability
            ws.cell(row=row, column=3, value=data['price'])         # Price
            ws.cell(row=row, column=4, value=data['reviews'])       # Reviews
            ws.cell(row=row, column=5, value=data['rating'])        # Rating
            ws.cell(row=row, column=6, value=data['electronics_rank']) # Electronics Rank
            ws.cell(row=row, column=7, value=data['last_updated'])  # Last Updated
            ws.cell(row=row, column=8, value=data['status'])        # Status
            
            # Color coding for out of stock (like Google Sheets)
            if data['availability'] == "Out of Stock":
                fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                ws.cell(row=row, column=2).fill = fill
            
            # Save the file
            wb.save(self.output_file)
            
        except Exception as e:
            self.logger.error(f"❌ Error updating output file: {str(e)}")
    
    def run_scraper(self):
        """Main function to run the scraper"""
        try:
            print("\n" + "="*60)
            print("🕷️  AMAZON SCRAPER STARTED")
            print("="*60)
            
            # Validate input file
            df, url_column = self.validate_input_file()
            
            # Create output file
            output_df = self.create_output_file(df, url_column)
            
            total_rows = len(df)
            processed = 0
            success_count = 0
            error_count = 0
            
            print(f"\n📊 Processing {total_rows} products...")
            print(f"🔄 Max retries per product: {self.max_retries}")
            print(f"⚙️ Processing mode: Sequential (one by one)")
            print("-" * 60)
            
            # Process products one by one sequentially
            for i in range(total_rows):
                url = df.iloc[i][url_column]
                processed += 1
                
                print(f"\n🔍 [{processed}/{total_rows}] Processing: {url}")
                
                # Scrape the product
                result = self.scrape_single_product(url)
                result['url'] = url
                
                # Update output file immediately
                self.update_output_file(i, result)
                
                # Update counters and show result
                if result['status'] == 'Success':
                    success_count += 1
                    print(f"✅ Success! Price: {result['price']}, Stock: {result['availability']}")
                else:
                    error_count += 1
                    print(f"❌ Error: {result['status']}")
                
                # Show progress
                progress_percent = (processed / total_rows) * 100
                print(f"📈 Progress: {processed}/{total_rows} ({progress_percent:.1f}%)")
                
                # Small delay between requests to avoid rate limiting
                if processed < total_rows:  # Don't delay after the last item
                    print("⏳ Waiting 2 seconds before next product...")
                    #time.sleep(2)
            
            # Final summary
            print("\n" + "="*60)
            print("🎉 SCRAPING COMPLETED!")
            print("="*60)
            print(f"📊 Total processed: {processed}")
            print(f"✅ Successful: {success_count}")
            print(f"❌ Errors: {error_count}")
            print(f"📁 Results saved to: {self.output_file}")
            print(f"📄 Log saved to: scraper.log")
            print("="*60)
            
            return True
            
        except Exception as e:
            self.logger.error(f"❌ Fatal error: {str(e)}")
            print(f"\n❌ SCRAPING FAILED: {str(e)}")
            return False

def main():
    """Main function to run the scraper"""
    try:
        # Initialize scraper
        scraper = AmazonScraper()
        
        # Run the scraper
        success = scraper.run_scraper()
        
        if success:
            print("\n✅ All done! Check your output.xlsx file.")
        else:
            print("\n❌ Scraping failed. Check the error messages above.")
            
    except Exception as e:
        print(f"\n💥 Critical error: {str(e)}")
        print("Please check if:")
        print("1. input.xlsx file exists in the same folder")
        print("2. The Excel file has a URL column")
        print("3. You have internet connection")

if __name__ == "__main__":
    main()
